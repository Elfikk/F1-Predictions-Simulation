import openpyxl as px

class PredictionPlayer():

    def __init__(self, file_path = "PredictionSpreadsheets/Jarek.xlsx"):

        self.preds = {}
        self.load_data_2024(file_path)

    def load_data_2024(self, file_path):
        wb = px.open(file_path, True)

        quali_h2h_sheet = wb["DriverQualiH2H"]
        quali_h2h = []
        for i in range(10):
            cell = quali_h2h_sheet.cell(2 + i, 3)
            quali_h2h.append(cell.internal_value)
        self.preds["QH2H"] = quali_h2h

        race_h2h_sheet = wb["DriverRaceH2H"]
        race_h2h = []
        for i in range(10):
            cell = race_h2h_sheet.cell(2 + i, 3)
            race_h2h.append(cell.internal_value)
        self.preds["RH2H"] = race_h2h

        constructor_sheet = wb["ConstructorPredictions"]
        constructor_preds = {}
        for i in range(10):
            team = constructor_sheet.cell(2+i, 1).internal_value
            pos = constructor_sheet.cell(2+i, 3).internal_value
            constructor_preds[team] = pos
        self.preds["CChamp"] = constructor_preds

        driver_sheet = wb["DriverPredictions"]
        driver_preds = {}
        for i in range(20):
            driver = driver_sheet.cell(2 + i, 1).internal_value
            pos = driver_sheet.cell(2 + i, 3).internal_value
            driver_preds[driver] = pos
        self.preds["DChamp"] = driver_preds

        naftern_sheet = wb["First6Races"]
        naftern = {}
        for i in range(6):
            driver_name = naftern_sheet.cell(3, i+2).internal_value
            naftern[i + 1] = driver_name
        self.preds["NAfterN"] = naftern

        bool_sheet = wb["WillThey"]
        podiums, poles, fls, q1s, monaco = {}, {}, {}, {}, {}
        for i in range(20):
            driver = bool_sheet.cell(2 + i, 1).internal_value
            podium = bool_sheet.cell(2 + i, 3).internal_value
            pole = bool_sheet.cell(2 + i, 5).internal_value
            fl = bool_sheet.cell(2 + i, 7).internal_value
            q1 = bool_sheet.cell(2 + i, 9).internal_value
            monago = bool_sheet.cell(2 + i, 11).internal_value

            podiums[driver] = podium
            poles[driver] = pole
            fls[driver] = fl
            q1s[driver] = q1
            monaco[driver] = monago

        self.preds["Podiums"] = podiums
        self.preds["Poles"] = poles
        self.preds["FLs"] = fls
        self.preds["Q1"] = q1s
        self.preds["Monaco"] = monaco

        superlative_sheet = wb["TheSuperlatives"]

        # Infinite sadness
        # Superlatives
        rows = list(range(3, 9)) + list(range(10, 13))
        keys = [
                "MaxDNF",
                "MinLaps",
                "DeltaPos",
                "DeltaPts",
                "MaxPS",
                "SlowStarter",
                "EngineComps",
                "FastestPS",
                "QCons",
                ]

        for i in range(len(keys)):
            self.preds[keys[i]] = superlative_sheet.cell(rows[i], 2).internal_value

        race_pick_sheet = wb["Pick5Races"]

        gasly_races = []
        hulk_races = []
        dnf_races = []

        for i in range(5):
            gasly_race = race_pick_sheet.cell(3, i + 1).internal_value
            gasly_races.append(gasly_race)

            hulk_race = race_pick_sheet.cell(8, i + 1).internal_value
            hulk_races.append(hulk_race)

            dnf_race = race_pick_sheet.cell(13, i + 1).internal_value
            dnf_races.append(dnf_race)

        self.preds["Hulk"] = hulk_races
        self.preds["Gasly"] = gasly_races
        self.preds["BlowyEngines"] = dnf_races

if __name__ == "__main__":

    Jarek = PredictionPlayer("PredictionSpreadsheets/Kacper.xlsx")

    # print(Jarek.preds)

    for key in Jarek.preds:
        print(key)
        print(Jarek.preds[key])
        print("")